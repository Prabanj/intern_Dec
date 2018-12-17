# GUI with clear (Contains Excel)

import serial
import tkinter

from PIL import Image, ImageTk
from tkinter import Tk, Label, BOTH


from openpyxl import Workbook

#global i
#i=0

wb = Workbook()
dest_filename = 'Data_log.xlsx'
#ws1 = wb.create_sheet(title="Dec")
sheet = wb.active
sheet.column_dimensions['A'].width = 30
sheet.cell(row=1, column=1).value = "Decoded value"



class simpleapp_tk(tkinter.Tk):
    
    
    def __init__(self,parent):
        tkinter.Tk.__init__(self,parent)
        self.parent = parent
        self.current_row=1
        self.initialize()
        self.maxsize(width=50000, height=50000)
        
        

    def initialize(self):

        self.grid()
        
        self.img = ImageTk.PhotoImage(Image.open('logo.jpg'))
        self.panel = tkinter.Label(image = self.img)
        self.panel.grid(row=0, column=0, columnspan=4)
        
        self.entryVariable1 = tkinter.StringVar()
        self.entry = tkinter.Entry(self,textvariable=self.entryVariable1)
        self.entry.grid(row=3, column=0, columnspan=2)
        
        
        
        button = tkinter.Button(self,text=u"Scan",command=self.OnButtonClick)
        button.grid(row=3, column=2)
        button2 = tkinter.Button(self,text=u"Clear",command=self.OnButtonClick2)
        button2.grid(row=3, column=3)
     
        
        self.labelVariable2 = tkinter.StringVar()
        label = tkinter.Label(self,textvariable=self.labelVariable2,anchor="w",font="Corbel 15",fg="black", bg="white",height =0, bd = 4, justify="center")
        label.grid(row=2,column=0,columnspan =5)
        self.labelVariable2.set(u"Backcase Number")
        
        self.labelVariable3 = tkinter.StringVar()
        label = tkinter.Label(self,textvariable=self.labelVariable3,anchor="w",fg="red",bg="white",font="Consolas 15",height=0, bd=4, justify="center")
        label.grid(row=5,column=0,columnspan=5)

        
    def OnButtonClick2(self):

        self.entry.focus_set()
        self.entry.selection_range(0, tkinter.END)
        self.entry.delete(0,"end")
       
        
    def OnButtonClick(self):
        self.s= serial.Serial()
        self.s.port = 'COM10'
        self.s.open()
        self.a=self.s.readline()
        self.entryVariable1.set(self.a)
        self.labelVariable3.set(self.a)
        self.a = (self.a[0:17])
        self.current_row +=1
        sheet.cell(row=self.current_row , column=1).value = self.a
#        self.B = str('B'+str(i+1))
#        ws1[self.B] = self.a
        wb.save(filename = dest_filename)


if __name__ == "__main__":  
    
    app = simpleapp_tk(None)
    app.title('UID Scanner')
    app.configure(bg='#fff')
    app.mainloop()